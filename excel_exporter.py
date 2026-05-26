"""
Excel export — single responsibility module.
Converts a Job's OCR result into a formatted .xlsx workbook.
"""

import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------------------
# Style constants (module-level — created once, shared across all calls)
# ---------------------------------------------------------------------------

_HDR_FILL   = PatternFill("solid", fgColor="2D6A4F")
_SUB_FILL   = PatternFill("solid", fgColor="52B788")
_ALT_FILL   = PatternFill("solid", fgColor="F0F7F2")
_AMBER_FILL = PatternFill("solid", fgColor="FFF3E0")
_AMBER_HDR_FILL = PatternFill("solid", fgColor="F4A261")

_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_SUB_FONT   = Font(bold=True, color="FFFFFF", size=9)
_BODY_FONT  = Font(size=10)
_AMBER_FONT = Font(bold=True, color="F4A261", size=10)
_OK_FONT    = Font(italic=True, color="2D6A4F", size=10)

_THIN   = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left", vertical="center")

_COL_LETTER = openpyxl.utils.get_column_letter   # callable ref, no re-import


def _style(cell, font=_BODY_FONT, fill=None, align=_CENTRE, border=_BORDER):
    cell.font      = font
    cell.alignment = align
    cell.border    = border
    if fill:
        cell.fill = fill


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def make_excel(job) -> io.BytesIO:
    """
    Build an .xlsx workbook from a Job and return it as a BytesIO buffer.

    Sheets produced:
      1. Summary          — document meta + flagged count
      2. Flagged for Review — every cell containing '?' (amber highlight)
      3. Page N …         — one sheet per scanned page
    """
    data     = job.get_data()
    filename = job.filename
    pages    = data.get("pages", [])

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.append(["Field", "Value"])
    for cell in ws_sum[1]:
        _style(cell, font=_HDR_FONT, fill=_HDR_FILL)

    summary = data.get("summary", {})
    ws_sum.append(["Source File",    filename])
    ws_sum.append(["Total Pages",    len(pages)])
    ws_sum.append(["Total Records",  len(data.get("all_pickers", []))])

    for key, val in summary.items():
        if key == "additional_fields" and isinstance(val, dict):
            for k2, v2 in val.items():
                ws_sum.append([k2.replace("_", " ").title(), v2])
        elif val:
            ws_sum.append([key.replace("_", " ").title(), val])

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 42

    # ── Sheet 2: Flagged for Review ───────────────────────────────────────────
    ws_flag = wb.create_sheet(title="Flagged for Review")
    ws_flag.append(["Page", "Row #", "Name / Label", "Column", "Extracted Value"])
    for cell in ws_flag[1]:
        _style(cell, font=_HDR_FONT, fill=_AMBER_HDR_FILL)

    flagged_count = 0
    for page in pages:
        pg_num  = page.get("page_number", 1)
        pg_rows = page.get("rows", page.get("pickers", []))
        pg_hdrs = page.get("column_headers", [])
        for ri, row in enumerate(pg_rows, 1):
            cell_data = row.get("fields") or row.get("quantities") or {}
            label     = row.get("name") or row.get("row_label") or ""
            for col_key in pg_hdrs:
                val = cell_data.get(col_key)
                if val is not None and "?" in str(val):
                    ws_flag.append(
                        [f"p{pg_num}", ri, label, col_key.replace("_", " "), val]
                    )
                    flag_row = ws_flag.max_row
                    for ci, cell in enumerate(ws_flag[flag_row], 1):
                        _style(cell,
                               font=_AMBER_FONT if ci == 5 else _BODY_FONT,
                               fill=_AMBER_FILL,
                               align=_LEFT)
                    flagged_count += 1

    if flagged_count == 0:
        ws_flag.append(["—", "—", "—", "—", "No uncertain readings ✓"])
        for cell in ws_flag[ws_flag.max_row]:
            _style(cell, font=_OK_FONT, align=_LEFT)

    ws_flag.column_dimensions["A"].width = 7
    ws_flag.column_dimensions["B"].width = 7
    ws_flag.column_dimensions["C"].width = 20
    ws_flag.column_dimensions["D"].width = 32
    ws_flag.column_dimensions["E"].width = 22

    # Append flagged count to Summary
    ws_sum.append(["Flagged for Review", flagged_count])

    # ── Sheets 3+: One per page ───────────────────────────────────────────────
    for page in pages:
        pg_num   = page.get("page_number", 1)
        rows     = page.get("rows", page.get("pickers", []))
        headers  = page.get("column_headers", [])
        doc_info = page.get("document_info", {})

        ws = wb.create_sheet(title=f"Page {pg_num}")

        # Doc-info block
        _DOC_FIELD_MAP = {
            "title": "Title", "doc_ref": "Doc Ref", "date": "Date",
            "day_no": "Day No", "location": "Location", "supervisor": "Supervisor",
        }
        info_written = 0
        for key, label in _DOC_FIELD_MAP.items():
            val = doc_info.get(key)
            if val:
                ws.append([label, val])
                info_written += 1
        for k, v in (doc_info.get("additional_fields") or {}).items():
            if v:
                ws.append([k.replace("_", " ").title(), v])
                info_written += 1
        if info_written == 0:
            ws.append(["Page", pg_num])
        ws.append([])   # blank separator

        # Determine which optional meta columns have data
        has_team      = any(r.get("team")                    for r in rows)
        has_name      = any(r.get("name") or r.get("row_label") for r in rows)
        has_record_id = any(r.get("record_id")               for r in rows)
        has_start     = any(r.get("start_time")              for r in rows)
        has_end       = any(r.get("end_time")                for r in rows)
        has_notes     = any(r.get("notes")                   for r in rows)

        fixed_cols = ["#", "Page"]
        if has_team:      fixed_cols.append("Team")
        if has_name:      fixed_cols.append("Name / Label")
        if has_record_id: fixed_cols.append("Record ID")
        if has_start:     fixed_cols.append("Start Time")

        tail_cols = []
        if has_end:   tail_cols.append("End Time")
        if has_notes: tail_cols.append("Notes")

        all_cols  = fixed_cols + headers + tail_cols
        qty_start = len(fixed_cols) + 1               # 1-based, constant per page
        qty_end   = len(fixed_cols) + len(headers)

        # Header row
        hdr_row = ws.max_row + 1
        ws.append(all_cols)
        for ci, cell in enumerate(ws[hdr_row], 1):
            if ci <= len(fixed_cols) or ci > qty_end:
                _style(cell, font=_HDR_FONT, fill=_HDR_FILL)
            else:
                _style(cell, font=_SUB_FONT, fill=_SUB_FILL)

        # Data rows
        for ri, p in enumerate(rows, 1):
            row_fill  = _ALT_FILL if ri % 2 == 0 else None
            cell_data = p.get("fields") or p.get("quantities") or {}

            row_vals = [ri, p.get("_page", "")]
            if has_team:      row_vals.append(p.get("team") or "")
            if has_name:      row_vals.append(p.get("name") or p.get("row_label") or "")
            if has_record_id: row_vals.append(p.get("record_id") or p.get("picker_no") or "")
            if has_start:     row_vals.append(p.get("start_time") or "")

            qty_vals = [
                cell_data.get(h) if cell_data.get(h) is not None else "—"
                for h in headers
            ]

            tail_vals = []
            if has_end:   tail_vals.append(p.get("end_time") or "")
            if has_notes: tail_vals.append(p.get("notes") or "")

            ws.append(row_vals + qty_vals + tail_vals)

            data_row = ws.max_row
            # name/label column index (3rd fixed col if present)
            name_ci = 3 if len(fixed_cols) >= 3 else None
            for ci, cell in enumerate(ws[data_row], 1):
                is_qty       = qty_start <= ci <= qty_end
                is_ambiguous = (
                    is_qty and cell.value is not None and "?" in str(cell.value)
                )
                _style(cell,
                       font=_AMBER_FONT if is_ambiguous else _BODY_FONT,
                       fill=row_fill,
                       align=_LEFT if ci == name_ci else _CENTRE)

        # Column widths
        ws.column_dimensions["A"].width = 4    # #
        ws.column_dimensions["B"].width = 6    # Page
        for ci in range(3, len(fixed_cols) + 1):
            ws.column_dimensions[_COL_LETTER(ci)].width = 16
        for ci, hdr in enumerate(headers, len(fixed_cols) + 1):
            ws.column_dimensions[_COL_LETTER(ci)].width = max(
                10, min(len(hdr.replace("_", " ")) + 2, 28)
            )
        for ci in range(len(fixed_cols) + len(headers) + 1, len(all_cols) + 1):
            ws.column_dimensions[_COL_LETTER(ci)].width = 14

        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=len(fixed_cols) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
